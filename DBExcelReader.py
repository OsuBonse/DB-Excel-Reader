from openpyxl import *
import sqlite3

workbook = load_workbook(r'F:\Prog\Py\b1.xlsx')
sheet = workbook['123']
sqlite_connection = sqlite3.connect('ltx.db')
cursor = sqlite_connection.cursor()
sqlite_insert_query = """INSERT INTO ltx (Id,Cudnum,Koatyy,Area,District,Settlement,Street,TGName,PurposeOfTheAssignment,NameOfTheSite,TransactionType,Price,OwnershipType,RegistrationDate,RegistrationNumber,ValueNGO,EvaluationDate) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)"""
for i in range(4, 33):
     a = sheet.cell(row=i, column = 2).value
     b = sheet.cell(row=i, column = 3).value
     c = sheet.cell(row=i, column = 4).value
     d = sheet.cell(row=i, column = 5).value
     e = sheet.cell(row=i, column = 6).value
     f = sheet.cell(row=i, column = 7).value
     g = sheet.cell(row=i, column = 8).value
     h = sheet.cell(row=i, column = 9).value
     x = sheet.cell(row=i, column = 10).value
     j = sheet.cell(row=i, column = 11).value
     k = sheet.cell(row=i, column = 12).value
     l = sheet.cell(row=i, column = 13).value
     m = sheet.cell(row=i, column = 14).value
     n = sheet.cell(row=i, column = 15).value
     o = sheet.cell(row=i, column = 16).value
     p = sheet.cell(row=i, column = 17).value
     q = sheet.cell(row=i, column = 18).value
     r = sheet.cell(row=i, column = 19).value
     cursor.execute(sqlite_insert_query,(a,b,c,d,e,f,g,h,x,j,k,l,m,n,o,p,q))
     sqlite_connection.commit()
cursor.close() 
sqlite_connection.close()
